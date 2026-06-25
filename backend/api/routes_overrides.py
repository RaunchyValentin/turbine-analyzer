import io
import json
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from database import get_db
from models import Turbine
from models.parameter import Parameter
from models.override import UserOverride

router = APIRouter(tags=["overrides"])


@router.get("/overrides")
async def list_overrides(turbine_id: int = Query(...), db: AsyncSession = Depends(get_db)):
    """Return all user overrides for a turbine with full parameter metadata."""
    res = await db.execute(
        select(UserOverride, Parameter).join(
            Parameter, UserOverride.parameter_id == Parameter.id
        ).where(UserOverride.turbine_id == turbine_id)
        .order_by(Parameter.name)
    )
    rows = []
    for ov, p in res.all():
        rd = {}
        if p.raw_data:
            try:
                rd = json.loads(p.raw_data)
            except Exception:
                pass
        rows.append({
            "id":             ov.id,
            "parameter_id":   p.id,
            "tag_name":       rd.get("Tag-Name", ""),
            "port_name":      rd.get("Port-Name", ""),
            "port_id":        rd.get("Port-ID", ""),
            "parameter_key":  p.kks or "",
            "original_value": ov.original_value or p.value or "",
            "user_value":     ov.user_value,
            "unit":           p.unit or rd.get("EU", ""),
            "designation":    p.description or rd.get("Designation", ""),
            "modified_at":    ov.modified_at.isoformat() if ov.modified_at else "",
            "note":           ov.note or "",
        })
    return rows


@router.delete("/overrides/{override_id}", status_code=204)
async def revert_override(override_id: int, db: AsyncSession = Depends(get_db)):
    """Revert a single override (does not update curve_points automatically)."""
    ov = await db.get(UserOverride, override_id)
    if not ov:
        raise HTTPException(404)
    await db.delete(ov)
    await db.commit()


@router.delete("/overrides", status_code=204)
async def revert_all_overrides(turbine_id: int = Query(...), db: AsyncSession = Depends(get_db)):
    """Revert all overrides for a turbine."""
    res = await db.execute(
        select(UserOverride).where(UserOverride.turbine_id == turbine_id)
    )
    for ov in res.scalars().all():
        await db.delete(ov)
    await db.commit()


@router.get("/export/overrides")
async def export_overrides(turbine_id: int = Query(...), db: AsyncSession = Depends(get_db)):
    """Export user overrides as SREL-like Excel file (only changed parameters)."""
    turbine = await db.get(Turbine, turbine_id)
    if not turbine:
        raise HTTPException(404, "Turbine not found")

    res = await db.execute(
        select(UserOverride, Parameter).join(
            Parameter, UserOverride.parameter_id == Parameter.id
        ).where(UserOverride.turbine_id == turbine_id)
        .order_by(Parameter.name)
    )

    rows = []
    for ov, p in res.all():
        rd = {}
        if p.raw_data:
            try:
                rd = json.loads(p.raw_data)
            except Exception:
                pass
        rows.append({
            "Diagram-Name":   rd.get("Diagram-Name", "") or p.group or "",
            "Tag-Name":       rd.get("Tag-Name", ""),
            "Port-Name":      rd.get("Port-Name", ""),
            "Port-ID":        rd.get("Port-ID", ""),
            "Parameter Key":  p.kks or "",
            "Original Value": ov.original_value or p.value or "",
            "New Value":      ov.user_value,
            "EU":             p.unit or rd.get("EU", ""),
            "Designation":    p.description or rd.get("Designation", ""),
            "Modified At":    ov.modified_at.strftime("%Y-%m-%d %H:%M") if ov.modified_at else "",
        })

    if not rows:
        raise HTTPException(404, "No overrides found for this turbine")

    xlsx = _build_overrides_excel(rows, turbine.name)
    filename = f"Overrides_{turbine.name}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _build_overrides_excel(rows: list[dict], turbine_name: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Overrides"

    COLS = ["Diagram-Name", "Tag-Name", "Port-Name", "Port-ID",
            "Parameter Key", "Original Value", "New Value",
            "EU", "Designation", "Modified At"]

    HEADER_FILL = PatternFill("solid", fgColor="1a3a5a")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    CHANGED_FILL = PatternFill("solid", fgColor="1a2a10")

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    title = ws.cell(1, 1, f"User Overrides — {turbine_name}  ({datetime.utcnow().strftime('%Y-%m-%d')})")
    title.font = Font(bold=True, size=12)
    title.alignment = Alignment(horizontal="left")

    # Header
    ws.append(COLS)
    for col in range(1, len(COLS) + 1):
        cell = ws.cell(2, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(c, "") for c in COLS])
        # Highlight New Value cell
        nv_cell = ws.cell(ws.max_row, COLS.index("New Value") + 1)
        nv_cell.fill = CHANGED_FILL
        nv_cell.font = Font(bold=True, color="88ff44")

    col_widths = [20, 22, 12, 8, 28, 14, 14, 6, 28, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{len(rows) + 2}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
