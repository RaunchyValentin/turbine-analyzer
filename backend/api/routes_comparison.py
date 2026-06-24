import json
import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from database import get_db
from models import Turbine, Parameter

router = APIRouter(tags=["comparison"])


def _match_key(p: dict) -> str:
    """Primary match key: clean kks, fallback to name."""
    k = (p.get("kks") or "").strip()
    return k if k else (p.get("name") or "")


def _is_jar(turbine: "Turbine") -> bool:
    return (turbine.source_file or "").lower().endswith(".jar")


def _srel_keys(params: list[dict]) -> set[str]:
    """Return the set of match keys from SREL parameters (have a real kks)."""
    return {_match_key(p) for p in params if (p.get("kks") or "").strip()}


def _compare(params_a: list[dict], params_b: list[dict],
             jar_a: bool, jar_b: bool) -> list[dict]:
    """
    Build comparison rows.

    JAR vs SREL (mixed): restrict to keys present in the SREL side.
    SREL vs SREL or JAR vs JAR: full union of keys.
    """
    idx_a = {_match_key(p): p for p in params_a if _match_key(p)}
    idx_b = {_match_key(p): p for p in params_b if _match_key(p)}

    if jar_a != jar_b:
        # Mixed: use SREL side as the key universe
        anchor_keys = _srel_keys(params_b if jar_a else params_a)
        all_keys = sorted(anchor_keys)
    else:
        all_keys = sorted(idx_a.keys() | idx_b.keys())

    return [_build_row(k, idx_a.get(k), idx_b.get(k)) for k in all_keys]


def _raw(p: dict) -> dict:
    rd = p.get("raw_data")
    if isinstance(rd, str):
        try:
            return json.loads(rd)
        except Exception:
            return {}
    return rd or {}


def _build_row(key: str, pa: dict | None, pb: dict | None) -> dict:
    src = pa or pb
    rd = _raw(src)
    tag  = rd.get("Tag-Name", "") or (src.get("name") or "").split("|")[0]
    port = rd.get("Port-Name", "") or ""
    desc = rd.get("Designation", "") or src.get("description", "") or ""
    pkey = rd.get("Parameter Key", "") or src.get("kks", "") or key
    eu   = rd.get("EU", "") or src.get("unit", "") or ""

    val_a = pa.get("value", "") if pa else None
    val_b = pb.get("value", "") if pb else None

    if pa is None:
        status = "only_b"
    elif pb is None:
        status = "only_a"
    elif str(val_a) == str(val_b):
        status = "matching"
    else:
        status = "changed"

    return {
        "key":       key,
        "tag_name":  tag,
        "port_name": port,
        "param_key": pkey,
        "description": desc,
        "eu":        eu,
        "value_a":   val_a,
        "value_b":   val_b,
        "status":    status,
    }


async def _load_params(db: AsyncSession, turbine_id: int) -> list[dict]:
    result = await db.execute(select(Parameter).where(Parameter.turbine_id == turbine_id))
    return [
        {c.name: getattr(p, c.name) for c in p.__table__.columns}
        for p in result.scalars().all()
    ]


@router.get("/comparison")
async def compare(turbine_a: int, turbine_b: int, db: AsyncSession = Depends(get_db)):
    ta = await db.get(Turbine, turbine_a)
    tb = await db.get(Turbine, turbine_b)
    if not ta or not tb:
        raise HTTPException(404, "Turbine not found")

    params_a = await _load_params(db, turbine_a)
    params_b = await _load_params(db, turbine_b)

    rows = _compare(params_a, params_b, _is_jar(ta), _is_jar(tb))

    stats = {"matching": 0, "changed": 0, "only_a": 0, "only_b": 0}
    for r in rows:
        stats[r["status"]] += 1

    mixed = _is_jar(ta) != _is_jar(tb)
    return {
        "turbine_a": {"id": ta.id, "name": ta.name, "file_date": ta.file_date, "source_file": ta.source_file},
        "turbine_b": {"id": tb.id, "name": tb.name, "file_date": tb.file_date, "source_file": tb.source_file},
        "stats": stats,
        "mixed": mixed,
        "rows": [r for r in rows if r["status"] != "matching"],
    }


@router.get("/comparison/export")
async def export_comparison(turbine_a: int, turbine_b: int, db: AsyncSession = Depends(get_db)):
    ta = await db.get(Turbine, turbine_a)
    tb = await db.get(Turbine, turbine_b)
    if not ta or not tb:
        raise HTTPException(404, "Turbine not found")

    params_a = await _load_params(db, turbine_a)
    params_b = await _load_params(db, turbine_b)

    rows = _compare(params_a, params_b, _is_jar(ta), _is_jar(tb))
    non_matching = [r for r in rows if r["status"] != "matching"]

    name_a = f"{ta.name} ({ta.file_date or ta.imported_at})"
    name_b = f"{tb.name} ({tb.file_date or tb.imported_at})"
    xlsx = _build_excel(non_matching, name_a, name_b)

    filename = f"comparison_{ta.name}_vs_{tb.name}.xlsx"
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _build_excel(rows: list[dict], name_a: str, name_b: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ORANGE = PatternFill("solid", fgColor="FFB347")
    YELLOW = PatternFill("solid", fgColor="FFF176")
    BLUE_A = PatternFill("solid", fgColor="BBDEFB")
    BLUE_B = PatternFill("solid", fgColor="C8E6C9")
    HEADER = PatternFill("solid", fgColor="263238")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    headers = ["#", "Tag-Name", "Parameter Key", "Port-Name", "Description", name_a, name_b, "EU", "Status"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    STATUS_LABELS = {"changed": "Changed", "only_a": f"Only in {name_a}", "only_b": f"Only in {name_b}"}

    for i, row in enumerate(rows, 1):
        ws.append([
            i,
            row["tag_name"],
            row["param_key"],
            row["port_name"],
            row["description"],
            row["value_a"] if row["value_a"] is not None else "—",
            row["value_b"] if row["value_b"] is not None else "—",
            row["eu"],
            STATUS_LABELS.get(row["status"], row["status"]),
        ])
        excel_row = i + 1
        if row["status"] == "changed":
            ws.cell(excel_row, 7).fill = ORANGE
        elif row["status"] == "only_a":
            ws.cell(excel_row, 6).fill = BLUE_A
            ws.cell(excel_row, 7).fill = BLUE_A
        elif row["status"] == "only_b":
            ws.cell(excel_row, 6).fill = BLUE_B
            ws.cell(excel_row, 7).fill = BLUE_B
        for col in range(1, 10):
            ws.cell(excel_row, col).border = border

    col_widths = [5, 28, 22, 14, 30, 16, 16, 8, 20]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
