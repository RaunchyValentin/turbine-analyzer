import io
import json
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from database import get_db
from models import Turbine, Parameter
from services import export_parameters_to_excel, export_comparison_to_excel, compare_turbines

router = APIRouter(tags=["export"])


def _extract_prefix(raw_key: str) -> str:
    """Strip leading special chars and return the part before ':', or '' if none."""
    bare = raw_key.lstrip("§$#@ ")
    colon = bare.find(":")
    return bare[:colon].strip().upper() if colon >= 0 else ""


@router.get("/export/key-types")
async def get_key_types(turbine_id: int = Query(...), db: AsyncSession = Depends(get_db)):
    """Return sorted list of unique Parameter Key prefixes for a turbine."""
    turbine = await db.get(Turbine, turbine_id)
    if not turbine:
        raise HTTPException(404, "Turbine not found")

    result = await db.execute(
        select(Parameter).where(Parameter.turbine_id == turbine_id)
    )
    counts: dict[str, int] = {}
    for p in result.scalars().all():
        try:
            rd = json.loads(p.raw_data) if isinstance(p.raw_data, str) else (p.raw_data or {})
        except Exception:
            rd = {}
        raw_key = rd.get("Parameter Key", "") or ""
        if not raw_key:
            continue
        pfx = _extract_prefix(raw_key)
        if pfx:
            counts[pfx] = counts.get(pfx, 0) + 1

    return [{"prefix": k, "count": v} for k, v in sorted(counts.items(), key=lambda x: -x[1])]


@router.get("/export/srel-from-jar")
async def export_srel_from_jar(
    turbine_id: int = Query(...),
    key_types: str = Query(""),
    key_filter: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    """Generate a SREL-format Excel from JAR-imported icdiagram.xml parameters."""
    turbine = await db.get(Turbine, turbine_id)
    if not turbine:
        raise HTTPException(404, "Turbine not found")

    result = await db.execute(
        select(Parameter).where(Parameter.turbine_id == turbine_id)
    )
    params = result.scalars().all()

    selected_types = {t.strip().upper() for t in key_types.split(",") if t.strip()}
    kf = key_filter.strip().lower()

    rows = []
    for p in params:
        rd_raw = p.raw_data
        rd = json.loads(rd_raw) if isinstance(rd_raw, str) else (rd_raw or {})
        raw_key = rd.get("Parameter Key", "") or ""

        # Filter by selected key type prefixes (if any selected)
        if selected_types:
            pfx = _extract_prefix(raw_key)
            if pfx not in selected_types:
                continue

        # Optional text filter: match against key, kks, designation, tag
        if kf:
            haystack = " ".join([
                raw_key.lower(),
                (p.kks or "").lower(),
                rd.get("Designation", "").lower(),
                rd.get("Tag-Name", "").lower(),
                rd.get("Diagram-Name", "").lower(),
            ])
            if kf not in haystack:
                continue

        rows.append({
            "Diagram-Name":  rd.get("Diagram-Name", "") or p.group or "",
            "Tag-Name":      rd.get("Tag-Name", "") or "",
            "Port-Name":     rd.get("Port-Name", "") or "",
            "Value":         p.value or "",
            "Parameter Key": raw_key,
            "EU":            rd.get("EU", "") or p.unit or "",
            "Designation":   rd.get("Designation", "") or p.description or "",
            "Signal Name":   rd.get("Signal Name", "") or "",
            "Variation min": "",
            "Variation max": "",
        })

    xlsx = _build_srel_excel(rows, turbine.name, turbine.file_date or str(turbine.imported_at or ""))
    filename = f"SREL_{turbine.name}_{turbine.file_date or 'export'}.xlsx"
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _build_srel_excel(rows: list[dict], turbine_name: str, file_date: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "SREL"

    COLS = ["Diagram-Name", "Tag-Name", "Port-Name", "Value",
            "Parameter Key", "EU", "Designation", "Signal Name",
            "Variation min", "Variation max"]

    HEADER_FILL = PatternFill("solid", fgColor="1565C0")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    title = ws.cell(1, 1, f"SREL Export — {turbine_name}  ({file_date})")
    title.font = Font(bold=True, size=12)
    title.alignment = Alignment(horizontal="left")

    # Header row
    ws.append(COLS)
    for col in range(1, len(COLS) + 1):
        cell = ws.cell(2, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(c, "") for c in COLS])

    col_widths = [22, 24, 14, 12, 28, 6, 28, 24, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{len(rows) + 2}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/export/parameters")
async def export_parameters(turbine_ids: list[int] = Query(...), db: AsyncSession = Depends(get_db)):
    turbines = []
    params_per_turbine = []

    for tid in turbine_ids:
        turbine = await db.get(Turbine, tid)
        if not turbine:
            raise HTTPException(404, f"Turbine {tid} not found")
        turbines.append({"name": turbine.name})
        result = await db.execute(select(Parameter).where(Parameter.turbine_id == tid))
        params = [
            {c.name: getattr(p, c.name) for c in p.__table__.columns}
            for p in result.scalars().all()
        ]
        params_per_turbine.append(params)

    xlsx_bytes = export_parameters_to_excel(turbines, params_per_turbine)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=parameters.xlsx"},
    )


@router.get("/export/comparison")
async def export_comparison(turbine_ids: list[int] = Query(...), db: AsyncSession = Depends(get_db)):
    turbine_names = []
    params_per_turbine = []

    for tid in turbine_ids:
        turbine = await db.get(Turbine, tid)
        if not turbine:
            raise HTTPException(404, f"Turbine {tid} not found")
        turbine_names.append(turbine.name)
        result = await db.execute(select(Parameter).where(Parameter.turbine_id == tid))
        params = [
            {c.name: getattr(p, c.name) for c in p.__table__.columns}
            for p in result.scalars().all()
        ]
        params_per_turbine.append(params)

    comparison_rows = compare_turbines(params_per_turbine)
    xlsx_bytes = export_comparison_to_excel(comparison_rows, turbine_names)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=comparison.xlsx"},
    )
